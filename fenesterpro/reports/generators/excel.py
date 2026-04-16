import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel(project):
    wb = openpyxl.Workbook()
    
    # Header styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Project Code", project.project_code])
    ws_summary.append(["Client", project.client_name])
    ws_summary.append(["Date", project.created_at.strftime("%Y-%m-%d")])
    ws_summary.append(["Status", project.status])
    
    # 2. Cut List Sheet
    ws_cut = wb.create_sheet(title="Cut List")
    headers = ["Profile Code", "Profile Name", "Window", "Piece Name", "Length (mm)", "Qty", "Total (mm)"]
    ws_cut.append(headers)
    for cell in ws_cut[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for w in project.windows.filter(computed=True):
        for cp in w.cut_pieces.all():
            ws_cut.append([
                cp.profile.code, cp.profile.name, w.item_code, 
                cp.piece_name, cp.length, cp.quantity * w.quantity, 
                cp.total_length
            ])
            
    # 3. Bar Optimisation Sheet
    ws_opt = wb.create_sheet(title="Bar Optimisation")
    if hasattr(project, 'optimisation_result'):
        data = project.optimisation_result.result_data
        
        headers = ["Profile Code", "Bar #", "Used (mm)", "Waste (mm)", "Waste %", "Pieces Diagram"]
        ws_opt.append(headers)
        for cell in ws_opt[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        for profile_code, p_data in data.get('profiles', {}).items():
            for bar in p_data.get('bar_layouts', []):
                pieces_str = " | ".join([f"{p['length']}mm" for p in bar['pieces']])
                waste_pct = (bar['remaining'] / 6000) * 100 if bar['remaining'] > 0 else 0
                ws_opt.append([
                    profile_code,
                    bar['bar_number'],
                    bar['used_length'],
                    bar['remaining'],
                    f"{waste_pct:.1f}%",
                    pieces_str
                ])
    else:
        ws_opt.append(["No optimisation data available."])

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    return wb
